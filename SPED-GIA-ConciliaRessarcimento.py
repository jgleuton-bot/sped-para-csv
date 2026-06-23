"""
SPED-GIA-ConciliaRessarcimento  -  Conciliacao do Ressarcimento de ICMS-ST
============================================================================
Cruza o ressarcimento de Substituicao Tributaria entre o SPED Fiscal e a GIA-SP,
por posto e por periodo, usando o CODIGO DO VISTO ELETRONICO como chave.

  - SPED Fiscal (Bloco E): registro E111 com COD_AJ_APUR = SP020799 (credito).
    O visto vem na DESCR_COMPL_AJ; o processo/Portaria CAT 42/2018 no E112.
  - GIA-SP (.prf): registro tipo 20, sub-item 007.99 (Outros creditos).
    Valor com 3 casas decimais (/1000); visto na descricao apos "n.".
    Periodo e CNPJ sao lidos do registro 05 do proprio .prf (nao dependem do
    nome da pasta). Havendo mais de uma GIA por posto/periodo (backup,
    substitutiva, retificada), usa-se a mais recente (data de modificacao).

A chave de cruzamento e o Codigo do Visto Eletronico, no mesmo posto (CNPJ) e
mesmo periodo. Para cada visto, o valor do SPED (SP020799) deve ser identico ao
da GIA (007.99).

Saidas (na pasta escolhida):
  - Conciliacao_SPEDxGIA.xlsx  -> aba Resumo, aba Metodo e 1 aba por posto
  - log_conciliacao.csv        -> auditoria (arquivos lidos / periodos)

Status por lancamento:
  OK ............... visto e valor identicos
  VISTO DIVERGENTE . valor confere, mas o Visto Eletronico difere
  VALOR DIVERGENTE . visto igual, valor diferente
  SO NO SPED ....... existe no SPED e nao na GIA
  SO NA GIA ........ existe na GIA e nao no SPED

Uso:
  python SPED-GIA-ConciliaRessarcimento.py
      -> abre janelas para escolher: pasta SPED, pasta GIA, pasta de saida
  python SPED-GIA-ConciliaRessarcimento.py C:/sped C:/gia C:/saida
      -> tudo por linha de comando

Requisitos: Python 3.x + openpyxl  (pip install openpyxl)
"""

__version__ = "1.0.0"

import csv, io, os, re, sys, time
from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

COD_AJ = "SP020799"          # ressarcimento ST no SPED (E111)
GIA_SUBITEM = "2000799"      # registro 20 + sub-item 007.99 na GIA

# --------------------------------------------------------------------------- #
# utilitarios
# --------------------------------------------------------------------------- #
def norm(v):
    return re.sub(r"\s+", " ", v or "").strip().upper()

def to_float(s):
    s = (s or "").strip()
    if not s:
        return 0.0
    try:
        return float(s.replace(".", "").replace(",", ".")) if "," in s else float(s)
    except ValueError:
        return 0.0

def per_label(ym):
    return f"{ym[4:6]}/{ym[0:4]}" if len(ym) == 6 else ym

def codnum(c):
    try:
        return int(re.sub(r"\D", "", c) or 99999)
    except ValueError:
        return 99999

# --------------------------------------------------------------------------- #
# SPED (worker em processo separado)
# --------------------------------------------------------------------------- #
def processar_sped(caminho):
    p = Path(caminho)
    res = {"arquivo": p.name, "cnpj": "", "nome": "", "periodo": "",
           "itens": [], "status": ""}
    try:
        raw = p.read_bytes()
        try:
            txt = raw.decode("utf-8")
        except UnicodeDecodeError:
            txt = raw.decode("latin-1")
        if not txt.startswith("|0000|"):
            res["status"] = "ignorado: nao e SPED"
            return res
        per_ini = ""
        for ln in txt.splitlines():
            if not ln.startswith("|"):
                continue
            f = ln.split("|")
            if len(f) < 2:
                continue
            reg = f[1]
            if reg == "0000":
                g = lambda i: f[i].strip() if i < len(f) else ""
                res["nome"] = g(6); res["cnpj"] = g(7)
                per_ini = g(4); res["periodo"] = g(4)[4:8] + g(4)[2:4]
            elif reg == "E100":
                di = f[2].strip() if len(f) > 2 else ""
                if len(di) == 8:
                    per_ini = di
            elif reg == "E111" and len(f) > 4 and f[2].strip() == COD_AJ:
                descr = f[3].strip()
                m = re.search(r"eletr\S*nico:?\s*(.*)$", descr, re.I)
                visto = m.group(1).strip() if m else ""
                ym = per_ini[4:8] + per_ini[2:4] if len(per_ini) == 8 else res["periodo"]
                res["itens"].append((ym, visto, to_float(f[4].strip())))
        res["status"] = "ok"
    except Exception as e:
        res["status"] = f"ERRO: {e}"
    return res

# --------------------------------------------------------------------------- #
# GIA (.prf)
# --------------------------------------------------------------------------- #
def parse_gia(caminho):
    """Retorna (cnpj, ym, [(visto, valor), ...]) ou None."""
    try:
        data = open(caminho, "rb").read().decode("latin-1", "replace")
    except Exception:
        return None
    linhas = data.splitlines()
    r05 = next((l for l in linhas if l.startswith("05")), "")
    if len(r05) < 30:
        return None
    cnpj = r05[14:28]
    m = re.match(r"0*1(20\d\d)(0[1-9]|1[0-2])", r05[28:])
    if not m or not cnpj.isdigit():
        return None
    ym = m.group(1) + m.group(2)
    ent = []
    for l in linhas:
        if l.startswith(GIA_SUBITEM):
            mv = re.match(r"(\d+)", l[7:])
            val = int(mv.group(1)) / 1000 if mv else 0.0
            i = l.find("n. ")
            visto = re.sub(r"\s{2,}.*$", "", l[i + 3:]).strip() if i >= 0 else ""
            ent.append((visto, val))
    return cnpj, ym, ent

_MESES = {"JANEIRO","FEVEREIRO","MARCO","MARÇO","ABRIL","MAIO","JUNHO","JULHO",
          "AGOSTO","SETEMBRO","OUTUBRO","NOVEMBRO","DEZEMBRO"}

def posto_de_pasta(caminho, raiz):
    """Extrai (cod, nome) da pasta do posto tipo '212 - ROSA BRANCA'.
    Ignora o nome do arquivo e as pastas de mes ('06 - JUNHO')."""
    rel = os.path.relpath(caminho, raiz)
    dirs = rel.split(os.sep)[:-1]                 # exclui o arquivo
    partes = [os.path.basename(raiz)] + dirs
    for parte in partes:                          # primeira pasta de posto
        m = re.match(r"^\s*(\d{2,4})\s*-\s*(.+?)\s*$", parte)
        if m and m.group(2).strip().upper() not in _MESES:
            return (m.group(1), m.group(2))
    return None

# --------------------------------------------------------------------------- #
# selecao de pastas
# --------------------------------------------------------------------------- #
def escolher():
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        a = input("Pasta do SPED (.txt): ").strip(' "')
        b = input("Pasta da GIA (.prf): ").strip(' "')
        c = input("Pasta de saida: ").strip(' "')
        if not (a and b and c):
            print("Entrada nao informada."); sys.exit(1)
        return Path(a), Path(b), Path(c)
    root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
    a = filedialog.askdirectory(parent=root, title="1/3  Pasta (e subpastas) onde esta o SPED Fiscal (.txt)")
    if not a:
        print("Cancelado."); sys.exit(1)
    b = filedialog.askdirectory(parent=root, title="2/3  Pasta (e subpastas) onde esta a GIA (.prf)", initialdir=a)
    if not b:
        print("Cancelado."); sys.exit(1)
    c = filedialog.askdirectory(parent=root, title="3/3  Pasta onde salvar os resultados", initialdir=b)
    if not c:
        print("Cancelado."); sys.exit(1)
    root.destroy()
    return Path(a), Path(b), Path(c)

# --------------------------------------------------------------------------- #
# reconciliacao
# --------------------------------------------------------------------------- #
def reconciliar(S, G, tem_gia):
    """S, G: listas de dicts {visto, val}. Retorna linhas + contadores."""
    S = [dict(x, used=False) for x in S]
    G = [dict(x, used=False) for x in G]
    rows = []
    for s in S:                                   # 1) por visto exato
        for g in G:
            if not g["used"] and s["visto"] and norm(g["visto"]) == norm(s["visto"]):
                s["used"] = g["used"] = True
                diff = round(s["val"] - g["val"], 2)
                rows.append([s["val"], g["val"], diff, s["visto"], g["visto"],
                             "OK" if abs(diff) < 0.005 else "VALOR DIVERGENTE"])
                break
    for s in S:                                   # 2) por valor
        if s["used"]:
            continue
        for g in G:
            if not g["used"] and abs(g["val"] - s["val"]) < 0.005:
                s["used"] = g["used"] = True
                rows.append([s["val"], g["val"], 0.0, s["visto"], g["visto"], "VISTO DIVERGENTE"])
                break
    for s in S:
        if not s["used"]:
            rows.append([s["val"], None, None, s["visto"], None,
                         "SO NO SPED" if tem_gia else "SO NO SPED (sem GIA no periodo)"])
    for g in G:
        if not g["used"]:
            rows.append([None, g["val"], None, None, g["visto"], "SO NA GIA"])
    return rows

# --------------------------------------------------------------------------- #
# XLSX
# --------------------------------------------------------------------------- #
def gerar_xlsx(postos, dados_sped, dados_gia, gia_src, labels, pasta_sai):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    F = "Arial"
    HF = PatternFill("solid", start_color="1F4E78"); HFONT = Font(name=F, bold=True, color="FFFFFF", size=10)
    CELL = Font(name=F, size=9); thin = Side(style="thin", color="D9D9D9"); BD = Border(thin, thin, thin, thin)
    OKF = PatternFill("solid", start_color="C6EFCE"); WARN = PatternFill("solid", start_color="FFEB9C")
    BAD = PatternFill("solid", start_color="FFC7CE"); PERF = PatternFill("solid", start_color="DDEBF7")
    TF = PatternFill("solid", start_color="1F4E78"); TFONT = Font(name=F, bold=True, size=12, color="FFFFFF")

    wb = Workbook(); wb.remove(wb.active)
    used = set()
    def sname(cod, nome):
        base = re.sub(r"[:\\/?*\[\]]", "", f"{cod}-{nome}")[:31]
        n = base or "Posto"; i = 1
        while n in used:
            suf = f"~{i}"; n = base[:31 - len(suf)] + suf; i += 1
        used.add(n); return n

    HDR = ["Valor SPED (R$)", "Valor GIA (R$)", "Diferenca", "Visto SPED", "Visto GIA", "Status"]
    summary = []
    for cnpj in postos:
        cod, nome = labels.get(cnpj, ("?", cnpj))
        ws = wb.create_sheet(sname(cod, nome)); ws.sheet_view.showGridLines = False
        for col in range(1, 7):
            ws.cell(row=1, column=col).fill = TF
        ws.cell(row=1, column=1, value=f"Posto {cod} - {nome}   |   CNPJ {cnpj}").font = TFONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.cell(row=2, column=1, value="Ressarcimento ST: SPED E111 SP020799 x GIA 007.99 (chave = Visto Eletronico)").font = Font(name=F, italic=True, size=9)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        row = 4
        yms = sorted({ym for (c, ym) in list(dados_sped) + list(dados_gia) if c == cnpj})
        tot_s = tot_g = 0.0; nok = nvd = nout = 0
        for ym in yms:
            S = dados_sped.get((cnpj, ym), []); G = dados_gia.get((cnpj, ym), [])
            if not S and not G:
                continue
            rr = reconciliar(S, G, (cnpj, ym) in dados_gia)
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PERF
            ws.cell(row=row, column=1, value=f"PERIODO {per_label(ym)}   (GIA: {gia_src.get((cnpj, ym), 'sem arquivo')})").font = Font(name=F, bold=True, size=10)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6); row += 1
            for i, h in enumerate(HDR, start=1):
                c = ws.cell(row=row, column=i, value=h); c.font = HFONT; c.fill = HF; c.border = BD
                c.alignment = Alignment(horizontal="center", wrap_text=True)
            row += 1
            sub_s = sub_g = 0.0
            for vs, vg, diff, ws_v, wg_v, st in rr:
                vals = [vs, vg, diff, ws_v, wg_v, st]
                for i, v in enumerate(vals, start=1):
                    c = ws.cell(row=row, column=i, value=v); c.font = CELL; c.border = BD
                    if i in (1, 2, 3) and v is not None:
                        c.number_format = "#,##0.00"
                ws.cell(row=row, column=6).fill = OKF if st == "OK" else (WARN if st == "VISTO DIVERGENTE" else BAD)
                if vs is not None: sub_s += vs
                if vg is not None: sub_g += vg
                nok += st == "OK"; nvd += st == "VISTO DIVERGENTE"; nout += st not in ("OK", "VISTO DIVERGENTE")
                row += 1
            ws.cell(row=row, column=4, value="Subtotal periodo").font = Font(name=F, bold=True, size=9)
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
            for col, val in ((1, sub_s), (2, sub_g), (3, round(sub_s - sub_g, 2))):
                c = ws.cell(row=row, column=col, value=round(val, 2)); c.font = Font(name=F, bold=True, size=9); c.number_format = "#,##0.00"; c.border = BD
            row += 2
            tot_s += sub_s; tot_g += sub_g
        for col, w in zip(range(1, 7), [16, 16, 12, 30, 30, 22]):
            ws.column_dimensions[get_column_letter(col)].width = w
        sit = "OK" if (nvd == 0 and nout == 0) else ("VISTO DIVERGENTE" if nout == 0 else "VERIFICAR")
        summary.append([cod, nome, cnpj, round(tot_s, 2), round(tot_g, 2), round(tot_s - tot_g, 2), nok, nvd, nout, sit])

    # Resumo
    wsr = wb.create_sheet("Resumo", 0)
    wsr.append(["Conciliacao Ressarcimento ST (SP020799 x GIA 007.99) - por posto"])
    wsr["A1"].font = Font(name=F, bold=True, size=13); wsr.append([])
    HH = ["Cod", "Posto", "CNPJ", "Soma SPED (R$)", "Soma GIA (R$)", "Diferenca", "# OK", "# Visto div.", "# Outros", "Situacao"]
    hr = wsr.max_row + 1; wsr.append(HH)
    for c in range(1, len(HH) + 1):
        x = wsr.cell(row=hr, column=c); x.fill = HF; x.font = HFONT; x.border = BD; x.alignment = Alignment(horizontal="center", wrap_text=True)
    for s in sorted(summary, key=lambda r: codnum(r[0])):
        wsr.append(s); r = wsr.max_row
        for c in range(1, len(HH) + 1):
            wsr.cell(row=r, column=c).font = CELL; wsr.cell(row=r, column=c).border = BD
        for c in (4, 5, 6):
            wsr.cell(row=r, column=c).number_format = "#,##0.00"
        sit = s[9]; wsr.cell(row=r, column=10).fill = OKF if sit == "OK" else (WARN if sit == "VISTO DIVERGENTE" else BAD)
    wsr.freeze_panes = f"A{hr + 1}"
    for col, w in zip("ABCDEFGHIJ", [7, 26, 16, 16, 16, 13, 7, 11, 9, 18]):
        wsr.column_dimensions[col].width = w

    # Metodo
    wsm = wb.create_sheet("Metodo")
    notas = [
        ("Como o cruzamento e feito", 13, True), ("", 10, False),
        ("Chave: Codigo do Visto Eletronico, no mesmo posto (CNPJ) e mesmo periodo.", 10, False),
        ("SPED: E111 com COD_AJ_APUR = SP020799 (visto na DESCR_COMPL_AJ).", 10, False),
        ("GIA (.prf): registro 20, sub-item 007.99; valor /1000; visto apos 'n.'.", 10, False),
        ("Periodo e CNPJ vem do registro 05 do .prf. GIA mais recente por posto/periodo.", 10, False),
        ("", 10, False),
        ("Status: OK | VISTO DIVERGENTE | VALOR DIVERGENTE | SO NO SPED | SO NA GIA.", 10, False),
        ("Priorize na aba Resumo as situacoes 'VERIFICAR' (diferenca de valor/quantidade).", 10, False),
    ]
    for i, (t, sz, b) in enumerate(notas, 1):
        wsm.cell(row=i, column=1, value=t).font = Font(name=F, size=sz, bold=b)
    wsm.column_dimensions["A"].width = 110

    path = Path(pasta_sai) / "Conciliacao_SPEDxGIA.xlsx"
    wb.save(path)
    return path, summary

# --------------------------------------------------------------------------- #
# principal
# --------------------------------------------------------------------------- #
def main():
    if len(sys.argv) > 3:
        psped, pgia, psai = Path(sys.argv[1]), Path(sys.argv[2]), Path(sys.argv[3])
    else:
        psped, pgia, psai = escolher()
    for d in (psped, pgia):
        if not d.exists():
            print(f"Pasta nao encontrada: {d}"); sys.exit(1)
    psai.mkdir(parents=True, exist_ok=True)
    try:
        import openpyxl  # noqa
    except ImportError:
        print("ERRO: instale o openpyxl  ->  pip install openpyxl"); sys.exit(1)

    print("=" * 70)
    print(f"  SPED-GIA-ConciliaRessarcimento  v{__version__}")
    print("=" * 70)
    arq_sped = sorted(str(p) for p in psped.rglob("*.txt"))
    arq_gia = sorted(str(p) for p in pgia.rglob("*.prf"))
    print(f"  SPED : {psped}   (.txt: {len(arq_sped)})")
    print(f"  GIA  : {pgia}   (.prf: {len(arq_gia)})")
    print(f"  Saida: {psai}")
    print("=" * 70 + "\n")

    t0 = time.time()
    workers = max(1, (os.cpu_count() or 2) - 1)

    # ---- SPED em paralelo ----
    dados_sped = defaultdict(list)   # (cnpj, ym) -> [{visto,val}]
    nome_por_cnpj = {}
    log = []
    print(f"Lendo SPED com {workers} processos...")
    feitos = 0; total = len(arq_sped)
    with ProcessPoolExecutor(max_workers=workers) as pool:
        futs = {pool.submit(processar_sped, a): a for a in arq_sped}
        for fut in as_completed(futs):
            r = fut.result(); feitos += 1
            log.append(["SPED", r["arquivo"], r["cnpj"], per_label(r["periodo"]), len(r["itens"]), r["status"]])
            if r["cnpj"] and r["nome"]:
                nome_por_cnpj.setdefault(r["cnpj"].zfill(14), r["nome"])
            for ym, visto, val in r["itens"]:
                dados_sped[(r["cnpj"].zfill(14), ym)].append({"visto": visto, "val": val})
            if feitos % 500 == 0 or feitos == total:
                print(f"  SPED {feitos}/{total}", flush=True)

    # ---- GIA (sequencial; arquivos pequenos) com dedup por mtime ----
    print("\nLendo GIA (.prf)...")
    cand = defaultdict(list)   # (cnpj, ym) -> [(mtime, entries, fname, path)]
    labels = {}                # cnpj -> (cod, nome)
    for f in arq_gia:
        r = parse_gia(f)
        if not r:
            continue
        cnpj, ym, ent = r; cnpj = cnpj.zfill(14)
        cand[(cnpj, ym)].append((os.path.getmtime(f), ent, os.path.basename(f)))
        lab = posto_de_pasta(f, str(pgia))
        if lab and cnpj not in labels:
            labels[cnpj] = lab
        log.append(["GIA", os.path.basename(f), cnpj, per_label(ym), len(ent), "ok"])
    dados_gia = {}; gia_src = {}
    for k, lst in cand.items():
        lst.sort(key=lambda x: x[0])
        mt, ent, fn = lst[-1]
        if not ent:
            com = [x for x in lst if x[1]]
            if com:
                mt, ent, fn = com[-1]
        if ent:
            dados_gia[k] = [{"visto": v, "val": val} for v, val in ent]
            gia_src[k] = fn

    # ---- labels finais (cod/posto) ----
    for cnpj in set([k[0] for k in dados_sped] + [k[0] for k in dados_gia]):
        if cnpj not in labels:
            labels[cnpj] = ("?", nome_por_cnpj.get(cnpj, cnpj))

    # ---- universo de postos com ressarcimento ----
    postos = sorted(
        {k[0] for k in dados_sped} | {k[0] for k in dados_gia},
        key=lambda c: codnum(labels.get(c, ("?",))[0]))
    postos = [c for c in postos if any(
        dados_sped.get((c, ym)) or dados_gia.get((c, ym))
        for ym in {ym for (cc, ym) in list(dados_sped) + list(dados_gia) if cc == c})]

    print(f"\nGerando Excel... ({len(postos)} postos com ressarcimento)")
    path, summary = gerar_xlsx(postos, dados_sped, dados_gia, gia_src, labels, psai)

    # log csv
    csv_log = Path(psai) / "log_conciliacao.csv"
    with open(csv_log, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(["Origem", "Arquivo", "CNPJ", "Periodo", "Qtd Ressarc.", "Status"])
        w.writerows(sorted(log, key=lambda x: (x[0], x[1])))

    ok = sum(1 for s in summary if s[9] == "OK")
    vd = sum(1 for s in summary if s[9] == "VISTO DIVERGENTE")
    vr = sum(1 for s in summary if s[9] == "VERIFICAR")
    print("\n" + "=" * 70)
    print(f"  Concluido em {time.time() - t0:.1f}s")
    print(f"  Postos: {len(summary)}   OK: {ok}   Visto divergente: {vd}   Verificar: {vr}")
    print(f"  -> {path}")
    print(f"  -> {csv_log}")
    print("=" * 70)


if __name__ == "__main__":
    import multiprocessing
    multiprocessing.freeze_support()
    main()
