"""
SPED-ExtratorBlocoE111  —  Extrator do Bloco E (registros E111 e filhos E112/E113)
====================================================================================
Varre uma pasta (e subpastas) com arquivos SPED Fiscal (EFD ICMS/IPI) .txt,
extrai os ajustes da apuracao do ICMS (registro E111) e seus registros-filho
(E112 e E113), identifica a empresa pelo CNPJ via planilha de parametros
(PostosR7.csv: Cod;Posto;CNPJ) e gera DOIS arquivos Excel:

  1) E111_Consolidado.xlsx
       - Resumo : totais gerais e por posto (formulas SUMIFS/COUNTIFS)
       - E111   : 1 linha por ajuste (com Cod/Posto/CNPJ, periodo e valor)
       - E112   : 1 linha por registro, ligada ao E111 pela coluna ID_E111
       - E113   : 1 linha por registro, ligada ao E111 pela coluna ID_E111

  2) E111_por_Posto.xlsx
       - Indice : lista de postos e suas abas
       - 1 aba por posto; dentro de cada aba os dados sao agrupados por periodo
         e, em cada periodo: bloco E111, abaixo o E112 e abaixo o E113.

Tambem grava log_E111.csv (auditoria: 1 linha por arquivo processado).

Regras:
  - Le o registro 0000 (cabecalho: empresa, CNPJ, IE, UF, periodo)
  - Coleta E111 (COD_AJ_APUR, DESCR_COMPL_AJ, VL_AJ_APUR) e os filhos E112/E113
  - Deduplica por (CNPJ + periodo), mantendo a retificadora mais recente
    (maior numero de sequencia no nome do arquivo; empate -> data de modificacao)
  - Detecta encoding automaticamente (UTF-8 ou Latin-1)
  - Nunca para por causa de um arquivo com problema: registra no log e continua

Uso:
    python SPED-ExtratorBlocoE111.py
        -> abre janelas para escolher: pasta dos SPEDs, arquivo PostosR7.csv, pasta de saida
    python SPED-ExtratorBlocoE111.py C:/speds C:/param/PostosR7.csv C:/saida
        -> tudo indicado por linha de comando

Requisitos: Python 3.x + openpyxl  (pip install openpyxl)
"""

__version__ = "1.0.0"

import csv
import io
import os
import re
import sys
import time
import datetime
from collections import defaultdict, Counter
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")


# --------------------------------------------------------------------------- #
# Conversoes
# --------------------------------------------------------------------------- #
def to_num(s):
    s = (s or "").strip()
    if s == "":
        return None
    try:
        return float(s.replace(".", "").replace(",", ".")) if "," in s else float(s)
    except ValueError:
        return s


def to_date(s):
    s = (s or "").strip()
    if len(s) == 8 and s.isdigit():
        try:
            return datetime.date(int(s[4:8]), int(s[2:4]), int(s[0:2]))
        except ValueError:
            return None
    return None


def per_sortkey(s):
    s = (s or "").strip()
    return s[4:8] + s[2:4] if len(s) == 8 else s


def per_label(s):
    s = (s or "").strip()
    return f"{s[2:4]}/{s[4:8]}" if len(s) == 8 else s


def seq_do_nome(nome):
    """Numero de sequencia/retificacao no nome: CNPJ-IE-DTINI-DTFIN-SEQ-HASH-..."""
    parts = nome.split("-")
    if len(parts) >= 5 and parts[4].isdigit():
        return int(parts[4])
    return 0


# --------------------------------------------------------------------------- #
# Worker: processa um arquivo SPED
# --------------------------------------------------------------------------- #
def processar_arquivo(caminho_str):
    p = Path(caminho_str)
    res = {
        "arquivo": str(p), "nome_arq": p.name, "mtime": 0.0,
        "seq": seq_do_nome(p.name),
        "cnpj": "", "ie": "", "nome": "", "uf": "",
        "dt_ini": "", "dt_fin": "", "ajustes": [], "status": "",
    }
    try:
        res["mtime"] = p.stat().st_mtime
        raw = p.read_bytes()
        try:
            texto = raw.decode("utf-8")
        except UnicodeDecodeError:
            texto = raw.decode("latin-1")

        if not texto.startswith("|0000|"):
            res["status"] = "ignorado: nao e arquivo SPED"
            return res

        per_ini = per_fin = ""
        atual = None  # E111 corrente
        for ln in texto.splitlines():
            if not ln.startswith("|"):
                continue
            f = ln.split("|")
            if len(f) < 2:
                continue
            reg = f[1]

            if reg == "0000":
                g = lambda i: f[i].strip() if i < len(f) else ""
                res["dt_ini"] = g(4); res["dt_fin"] = g(5)
                res["nome"] = g(6); res["cnpj"] = g(7)
                res["uf"] = g(9); res["ie"] = g(10)
                per_ini, per_fin = res["dt_ini"], res["dt_fin"]

            elif reg == "E100":
                g = lambda i: f[i].strip() if i < len(f) else ""
                per_ini, per_fin = g(2), g(3)

            elif reg == "E111":
                g = lambda i: f[i].strip() if i < len(f) else ""
                atual = {
                    "per_ini": per_ini or res["dt_ini"],
                    "per_fin": per_fin or res["dt_fin"],
                    "cod_aj": g(2), "descr": g(3), "valor": g(4),
                    "e112": [], "e113": [],
                }
                res["ajustes"].append(atual)

            elif reg == "E112" and atual is not None:
                g = lambda i: f[i].strip() if i < len(f) else ""
                atual["e112"].append(
                    [g(2), g(3), g(4), g(5), g(6)]  # NUM_DA,NUM_PROC,IND_PROC,PROC,TXT_COMPL
                )

            elif reg == "E113" and atual is not None:
                g = lambda i: f[i].strip() if i < len(f) else ""
                atual["e113"].append(
                    [g(2), g(3), g(4), g(5), g(6), g(7), g(8), g(9)]
                    # COD_PART,COD_MOD,SER,SUB,NUM_DOC,DT_DOC,COD_ITEM,VL_AJ_ITEM
                )

        res["status"] = "ok"
    except Exception as e:
        res["status"] = f"ERRO: {e}"
    return res


# --------------------------------------------------------------------------- #
# Selecao de entradas (GUI ou linha de comando)
# --------------------------------------------------------------------------- #
def escolher_entradas():
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        ent = input("Pasta-raiz com os SPEDs (.txt): ").strip(' "')
        csvp = input("Arquivo de parametros PostosR7.csv: ").strip(' "')
        sai = input("Pasta para salvar os XLSX: ").strip(' "')
        if not ent or not csvp or not sai:
            print("Entrada nao informada. Operacao cancelada.")
            sys.exit(1)
        return Path(ent), Path(csvp), Path(sai)

    root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
    ent = filedialog.askdirectory(parent=root, title="1/3  Selecione a pasta-raiz com os arquivos SPED (.txt)")
    if not ent:
        print("Nenhuma pasta selecionada. Operacao cancelada."); root.destroy(); sys.exit(1)
    csvp = filedialog.askopenfilename(
        parent=root, title="2/3  Selecione o arquivo de parametros (PostosR7.csv)",
        initialdir=ent, filetypes=[("CSV", "*.csv"), ("Todos", "*.*")])
    if not csvp:
        print("Nenhum CSV selecionado. Operacao cancelada."); root.destroy(); sys.exit(1)
    sai = filedialog.askdirectory(parent=root, title="3/3  Selecione a pasta onde salvar os XLSX", initialdir=ent)
    if not sai:
        print("Nenhuma pasta de saida selecionada. Operacao cancelada."); root.destroy(); sys.exit(1)
    root.destroy()
    return Path(ent), Path(csvp), Path(sai)


def carregar_mapa(csv_path):
    """PostosR7.csv (Cod;Posto;CNPJ) -> {cnpj14: {'cod','posto'}}."""
    m = {}
    with open(csv_path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            cnpj = (row.get("CNPJ") or "").strip().zfill(14)
            m[cnpj] = {"cod": (row.get("Cod") or "").strip(),
                       "posto": (row.get("Posto") or "").strip()}
    return m


def codnum(c):
    try:
        return int(c)
    except (ValueError, TypeError):
        return 99999


# --------------------------------------------------------------------------- #
# Geracao dos XLSX
# --------------------------------------------------------------------------- #
def gerar_xlsx(empresas, mapa, pasta_sai):
    """empresas: lista de dicts deduplicados (1 por CNPJ+periodo) com 'ajustes'."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    F = "Arial"
    HDR_FILL = PatternFill("solid", start_color="1F4E78")
    HDR_FONT = Font(name=F, bold=True, color="FFFFFF", size=10)
    CELL = Font(name=F, size=10)
    thin = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def info(cnpj):
        return mapa.get(cnpj.zfill(14), {"cod": "", "posto": ""})

    # ---- achatar em listas E111/E112/E113 com ID global ----
    e111_rows, e112_rows, e113_rows = [], [], []
    eid = 0
    for emp in empresas:
        ci = info(emp["cnpj"])
        for aj in emp["ajustes"]:
            eid += 1
            e111_rows.append({
                "id": eid, "cod": ci["cod"], "posto": ci["posto"],
                "cnpj": emp["cnpj"], "ie": emp["ie"], "nome": emp["nome"], "uf": emp["uf"],
                "per_ini": aj["per_ini"], "per_fin": aj["per_fin"],
                "cod_aj": aj["cod_aj"], "descr": aj["descr"], "valor": aj["valor"],
                "q112": len(aj["e112"]), "q113": len(aj["e113"]), "arq": emp["nome_arq"],
            })
            for x in aj["e112"]:
                e112_rows.append({"id": eid, "ci": ci, "cnpj": emp["cnpj"],
                                  "per": aj["per_ini"], "v": x, "arq": emp["nome_arq"]})
            for x in aj["e113"]:
                e113_rows.append({"id": eid, "ci": ci, "cnpj": emp["cnpj"],
                                  "per": aj["per_ini"], "v": x, "arq": emp["nome_arq"]})

    # ===================== MODELO 1: CONSOLIDADO ===================== #
    wb = Workbook()

    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
        ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 28

    def widths(ws, ws_widths):
        for i, w in enumerate(ws_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    ws = wb.active; ws.title = "E111"
    h = ["ID", "Cod", "Posto", "CNPJ", "IE", "Empresa", "UF", "Periodo", "Dt_Ini",
         "Dt_Fin", "COD_AJ_APUR", "DESCR_COMPL_AJ", "VL_AJ_APUR", "Qtd_E112", "Qtd_E113", "Arquivo"]
    ws.append(h)
    for r in e111_rows:
        ws.append([r["id"], r["cod"], r["posto"], r["cnpj"], r["ie"], r["nome"], r["uf"],
                   per_label(r["per_ini"]), to_date(r["per_ini"]), to_date(r["per_fin"]),
                   r["cod_aj"], r["descr"], to_num(r["valor"]), r["q112"], r["q113"], r["arq"]])
    nrow = len(e111_rows) + 1
    style_header(ws, len(h))
    for row in ws.iter_rows(min_row=2, max_row=nrow):
        for cell in row:
            cell.font = CELL; cell.border = BORDER
        row[8].number_format = "DD/MM/YYYY"; row[9].number_format = "DD/MM/YYYY"
        row[12].number_format = "#,##0.00"
    widths(ws, [6, 7, 20, 16, 15, 38, 5, 9, 11, 11, 13, 55, 14, 9, 9, 60])
    if nrow > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(h))}{nrow}"

    ws2 = wb.create_sheet("E112")
    h2 = ["ID_E111", "Cod", "Posto", "CNPJ", "Periodo", "NUM_DA", "NUM_PROC", "IND_PROC", "PROC", "TXT_COMPL", "Arquivo"]
    ws2.append(h2)
    for r in e112_rows:
        v = r["v"]
        ws2.append([r["id"], r["ci"]["cod"], r["ci"]["posto"], r["cnpj"], per_label(r["per"]),
                    v[0], v[1], v[2], v[3], v[4], r["arq"]])
    n2 = len(e112_rows) + 1; style_header(ws2, len(h2))
    for row in ws2.iter_rows(min_row=2, max_row=n2):
        for cell in row:
            cell.font = CELL; cell.border = BORDER
    widths(ws2, [8, 7, 20, 16, 9, 14, 16, 9, 40, 50, 60])
    if n2 > 1:
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(h2))}{n2}"

    ws3 = wb.create_sheet("E113")
    h3 = ["ID_E111", "Cod", "Posto", "CNPJ", "Periodo", "COD_PART", "COD_MOD", "SER", "SUB",
          "NUM_DOC", "DT_DOC", "COD_ITEM", "VL_AJ_ITEM", "Arquivo"]
    ws3.append(h3)
    for r in e113_rows:
        v = r["v"]
        ws3.append([r["id"], r["ci"]["cod"], r["ci"]["posto"], r["cnpj"], per_label(r["per"]),
                    v[0], v[1], v[2], v[3], v[4], to_date(v[5]), v[6], to_num(v[7]), r["arq"]])
    n3 = len(e113_rows) + 1; style_header(ws3, len(h3))
    for row in ws3.iter_rows(min_row=2, max_row=n3):
        for cell in row:
            cell.font = CELL; cell.border = BORDER
        row[10].number_format = "DD/MM/YYYY"; row[12].number_format = "#,##0.00"
    widths(ws3, [8, 7, 20, 16, 9, 14, 9, 6, 6, 12, 12, 16, 14, 60])
    if n3 > 1:
        ws3.auto_filter.ref = f"A1:{get_column_letter(len(h3))}{n3}"

    # Resumo
    wsr = wb.create_sheet("Resumo", 0)
    wsr.append(["Resumo dos Ajustes de Apuracao do ICMS (Registro E111)"])
    wsr["A1"].font = Font(name=F, bold=True, size=13)
    wsr.append([]); wsr.append(["Total de registros E111:", len(e111_rows)])
    wsr.append(["Total de registros E112:", len(e112_rows)])
    wsr.append(["Total de registros E113:", len(e113_rows)])
    wsr.append(["Soma VL_AJ_APUR (R$):", f"=SUM(E111!M2:M{nrow})" if nrow > 1 else 0])
    wsr["B6"].number_format = "#,##0.00"
    wsr.append([]); wsr.append(["Cod", "Posto", "CNPJ", "Qtd E111", "Soma VL_AJ_APUR (R$)"])
    hr = wsr.max_row
    seen, seset = [], set()
    for r in e111_rows:
        k = (r["cod"], r["posto"], r["cnpj"])
        if k not in seset:
            seset.add(k); seen.append(k)
    seen.sort(key=lambda k: codnum(k[0]))
    for cod, posto, cnpj in seen:
        rr = wsr.max_row + 1
        wsr.append([cod, posto, cnpj,
                    f"=COUNTIFS(E111!D2:D{nrow},C{rr})" if nrow > 1 else 0,
                    f"=SUMIFS(E111!M2:M{nrow},E111!D2:D{nrow},C{rr})" if nrow > 1 else 0])
        wsr.cell(row=rr, column=5).number_format = "#,##0.00"
    last = wsr.max_row
    for c in range(1, 6):
        cell = wsr.cell(row=hr, column=c); cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
    for row in wsr.iter_rows(min_row=hr + 1, max_row=last, min_col=1, max_col=5):
        for cell in row:
            cell.font = CELL; cell.border = BORDER
    for col, w in zip("ABCDE", [8, 22, 16, 10, 20]):
        wsr.column_dimensions[col].width = w
    wsr.freeze_panes = f"A{hr + 1}"

    path1 = Path(pasta_sai) / "E111_Consolidado.xlsx"
    wb.save(path1)

    # ===================== MODELO 2: POR POSTO ===================== #
    wb2 = Workbook(); wb2.remove(wb2.active)
    title_font = Font(name=F, bold=True, size=13, color="FFFFFF")
    title_fill = PatternFill("solid", start_color="1F4E78")
    per_font = Font(name=F, bold=True, size=11, color="FFFFFF")
    per_fill = PatternFill("solid", start_color="2E75B6")
    e111_fill = PatternFill("solid", start_color="C6E0B4")
    e112_fill = PatternFill("solid", start_color="FFE699")
    e113_fill = PatternFill("solid", start_color="BDD7EE")
    sec_font = Font(name=F, bold=True, size=10)
    hdr_font = Font(name=F, bold=True, size=9)
    hdr_fill = PatternFill("solid", start_color="D9D9D9")
    cell_font = Font(name=F, size=9)
    tot_font = Font(name=F, bold=True, size=9)
    NCOL = 8

    # agrupa empresas por CNPJ
    by_cnpj = defaultdict(list)
    for emp in empresas:
        if emp["ajustes"]:
            by_cnpj[emp["cnpj"].zfill(14)].append(emp)
    cnpjs = sorted(by_cnpj.keys(), key=lambda c: (codnum(info(c)["cod"]), c))

    used = set()
    def sheet_name(cod, posto):
        base = re.sub(r"[:\\/?*\[\]]", "", f"{cod}-{posto}")[:31]
        n = base or "Posto"; i = 1
        while n in used:
            suf = f"~{i}"; n = base[:31 - len(suf)] + suf; i += 1
        used.add(n); return n

    H111 = ["COD_AJ_APUR", "DESCR_COMPL_AJ", "VL_AJ_APUR"]
    H112 = ["NUM_DA", "NUM_PROC", "IND_PROC", "PROC", "TXT_COMPL"]
    H113 = ["COD_PART", "COD_MOD", "SER", "SUB", "NUM_DOC", "DT_DOC", "COD_ITEM", "VL_AJ_ITEM"]

    def banner(ws, row, text, fill):
        for col in range(1, NCOL + 1):
            ws.cell(row=row, column=col).fill = fill
        c = ws.cell(row=row, column=1, value=text); c.font = sec_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOL)
        c.alignment = Alignment(horizontal="left", vertical="center")
        return row + 1

    def thead(ws, row, headers):
        for i, hh in enumerate(headers, start=1):
            c = ws.cell(row=row, column=i, value=hh)
            c.font = hdr_font; c.fill = hdr_fill; c.border = BORDER
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        return row + 1

    sheet_index = []
    for c in cnpjs:
        ci = info(c); recs = by_cnpj[c]
        nome = recs[0]["nome"]; ie = recs[0]["ie"]
        sn = sheet_name(ci["cod"], ci["posto"])
        sheet_index.append((ci["cod"], ci["posto"], c, sn))
        ws = wb2.create_sheet(sn); ws.sheet_view.showGridLines = False
        for col in range(1, NCOL + 1):
            ws.cell(row=1, column=col).fill = title_fill
        ws.cell(row=1, column=1, value=f"Posto {ci['cod']} - {ci['posto']}").font = title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
        ws.cell(row=2, column=1, value=f"{nome}   |   CNPJ: {c}   |   IE: {ie}").font = Font(name=F, size=9, italic=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOL)
        ws.row_dimensions[1].height = 20

        # periodos -> ajustes
        pers = defaultdict(list)
        for emp in recs:
            for aj in emp["ajustes"]:
                pers[per_sortkey(aj["per_ini"])].append(aj)
        row = 4
        for pk in sorted(pers.keys()):
            ajs = pers[pk]
            for col in range(1, NCOL + 1):
                ws.cell(row=row, column=col).fill = per_fill
            ws.cell(row=row, column=1, value=f"PERIODO: {per_label(ajs[0]['per_ini'])}").font = per_font
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOL)
            ws.row_dimensions[row].height = 18; row += 1

            row = banner(ws, row, "E111 - Ajuste / Beneficio / Incentivo da Apuracao do ICMS", e111_fill)
            row = thead(ws, row, H111)
            first = row
            for aj in ajs:
                ws.cell(row=row, column=1, value=aj["cod_aj"]).font = cell_font
                ws.cell(row=row, column=2, value=aj["descr"]).font = cell_font
                vc = ws.cell(row=row, column=3, value=to_num(aj["valor"]))
                vc.font = cell_font; vc.number_format = "#,##0.00"
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = BORDER
                row += 1
            tl = ws.cell(row=row, column=2, value="Total do periodo")
            tl.font = tot_font; tl.alignment = Alignment(horizontal="right")
            tv = ws.cell(row=row, column=3, value=f"=SUM(C{first}:C{row-1})")
            tv.font = tot_font; tv.number_format = "#,##0.00"; tv.border = BORDER
            row += 1

            r112 = [x for aj in ajs for x in aj["e112"]]
            if r112:
                row = banner(ws, row, "E112 - Informacoes Adicionais dos Ajustes", e112_fill)
                row = thead(ws, row, H112)
                for v in r112:
                    for i, val in enumerate(v, start=1):
                        cc = ws.cell(row=row, column=i, value=val); cc.font = cell_font; cc.border = BORDER
                    row += 1
            r113 = [x for aj in ajs for x in aj["e113"]]
            if r113:
                row = banner(ws, row, "E113 - Identificacao dos Documentos Fiscais", e113_fill)
                row = thead(ws, row, H113)
                for v in r113:
                    vals = [v[0], v[1], v[2], v[3], v[4], to_date(v[5]), v[6], to_num(v[7])]
                    for i, val in enumerate(vals, start=1):
                        cc = ws.cell(row=row, column=i, value=val); cc.font = cell_font; cc.border = BORDER
                        if i == 6:
                            cc.number_format = "DD/MM/YYYY"
                        if i == 8:
                            cc.number_format = "#,##0.00"
                    row += 1
            row += 1
        for col, w in zip(range(1, NCOL + 1), [14, 46, 14, 12, 10, 12, 16, 14]):
            ws.column_dimensions[get_column_letter(col)].width = w

    idx = wb2.create_sheet("Indice", 0)
    idx.append(["Indice de Postos"]); idx["A1"].font = Font(name=F, bold=True, size=13)
    idx.append([]); idx.append(["Cod", "Posto", "CNPJ", "Aba"])
    for col in range(1, 5):
        cc = idx.cell(row=3, column=col); cc.font = Font(name=F, bold=True, color="FFFFFF")
        cc.fill = title_fill; cc.alignment = Alignment(horizontal="center")
    for cod, posto, cnpj, sn in sheet_index:
        idx.append([cod, posto, cnpj, sn])
    for col, w in zip("ABCD", [8, 24, 16, 28]):
        idx.column_dimensions[col].width = w
    idx.freeze_panes = "A4"

    path2 = Path(pasta_sai) / "E111_por_Posto.xlsx"
    wb2.save(path2)
    return path1, path2, len(e111_rows), len(e112_rows), len(e113_rows)


# --------------------------------------------------------------------------- #
# Principal
# --------------------------------------------------------------------------- #
def main():
    if len(sys.argv) > 3:
        pasta_ent, csv_param, pasta_sai = Path(sys.argv[1]), Path(sys.argv[2]), Path(sys.argv[3])
    else:
        pasta_ent, csv_param, pasta_sai = escolher_entradas()

    if not pasta_ent.exists():
        print(f"Pasta nao encontrada: {pasta_ent}"); sys.exit(1)
    if not csv_param.exists():
        print(f"CSV de parametros nao encontrado: {csv_param}"); sys.exit(1)
    pasta_sai.mkdir(parents=True, exist_ok=True)

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("ERRO: a biblioteca 'openpyxl' nao esta instalada.")
        print("Instale com:  pip install openpyxl")
        sys.exit(1)

    mapa = carregar_mapa(csv_param)

    print("=" * 70)
    print(f"  SPED-ExtratorBlocoE111  v{__version__}")
    print("=" * 70)
    print(f"  SPEDs     : {pasta_ent}  (inclui subpastas)")
    print(f"  Parametros: {csv_param}  ({len(mapa)} postos)")
    print(f"  Saida     : {pasta_sai}")

    arquivos = sorted(str(p) for p in pasta_ent.rglob("*.txt"))
    print(f"  Arquivos .txt encontrados: {len(arquivos)}")
    print("=" * 70 + "\n")
    if not arquivos:
        print("Nenhum arquivo .txt encontrado."); sys.exit(0)

    t0 = time.time()
    workers = max(1, (os.cpu_count() or 2) - 1)
    print(f"Processando com {workers} processos em paralelo...\n")

    # dedup por (cnpj, periodo) mantendo a retificadora mais recente
    melhor = {}     # (cnpj, dt_ini, dt_fin) -> res
    log = []
    feitos = 0; total = len(arquivos); larg = len(str(total))

    with ProcessPoolExecutor(max_workers=workers) as pool:
        futuros = {pool.submit(processar_arquivo, a): a for a in arquivos}
        for fut in as_completed(futuros):
            r = fut.result(); feitos += 1
            log.append([r["nome_arq"], r["nome"], r["cnpj"],
                        f"{per_label(r['dt_ini'])} a {per_label(r['dt_fin'])}",
                        len(r["ajustes"]), r["status"]])
            if r["status"] == "ok" and r["ajustes"]:
                ch = (r["cnpj"], r["dt_ini"], r["dt_fin"])
                ant = melhor.get(ch)
                if ant is None or (r["seq"], r["mtime"]) > (ant["seq"], ant["mtime"]):
                    melhor[ch] = r

            marca = "OK   " if r["status"] == "ok" else ("PULA " if r["status"].startswith("ignorado") else "ERRO ")
            pct = 100 * feitos / total
            emp = (r["nome"] or "-")[:35]
            linha = (f"  [{feitos:>{larg}}/{total}] {pct:5.1f}%  {marca} "
                     f"{emp:<35}  E111: {len(r['ajustes']):>3}  {r['nome_arq']}")
            if marca == "ERRO ":
                linha += f"  -> {r['status']}"
            print(linha, flush=True)

    empresas = list(melhor.values())

    # log de auditoria
    csv_log = Path(pasta_sai) / "log_E111.csv"
    log.sort(key=lambda x: x[0])
    with open(csv_log, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Arquivo", "Empresa", "CNPJ", "Periodo", "Qtd E111", "Status"])
        w.writerows(log)

    print("\nGerando arquivos Excel...")
    path1, path2, n111, n112, n113 = gerar_xlsx(empresas, mapa, pasta_sai)

    cnpjs_sem_mapa = sorted({e["cnpj"].zfill(14) for e in empresas
                             if e["cnpj"].zfill(14) not in mapa})
    erros = sum(1 for l in log if str(l[5]).startswith("ERRO"))
    ignorados = sum(1 for l in log if str(l[5]).startswith("ignorado"))

    print("\n" + "=" * 70)
    print(f"  Concluido em {time.time() - t0:.1f}s")
    print(f"  Arquivos OK: {len(log) - erros - ignorados}   Ignorados: {ignorados}   Erros: {erros}")
    print(f"  Periodos com E111 (apos dedup): {len(empresas)}")
    print(f"  Registros -> E111: {n111}   E112: {n112}   E113: {n113}")
    if cnpjs_sem_mapa:
        print(f"  [AVISO] {len(cnpjs_sem_mapa)} CNPJ(s) sem correspondencia no PostosR7.csv:")
        for c in cnpjs_sem_mapa[:10]:
            print(f"          {c}")
    print(f"  -> {path1}")
    print(f"  -> {path2}")
    print(f"  -> {csv_log}")
    print("=" * 70)


if __name__ == "__main__":
    import multiprocessing
    multiprocessing.freeze_support()
    main()
