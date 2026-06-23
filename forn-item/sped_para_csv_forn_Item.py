"""
Extrator de Notas e Itens de Entrada — SPED Fiscal -> CSV/XLSX  —  v2.1.0
===============================================================================
Varre uma pasta (e subpastas) com arquivos SPED .txt e, com base numa TABELA DE
FORNECEDORES SELECIONADOS informada antes do processamento, extrai:
  1) as notas fiscais de entrada (registro C100, modelo 55) desses fornecedores;
  2) os produtos/itens adquiridos (registro C170) de cada nota — com quantidade,
     descricao e valor total do item.

Identificacao do posto:
  As duas primeiras colunas das saidas sao "Cod" e "Posto", obtidas da tabela
  PostosR7.csv cruzando pelo CNPJ da Sociedade (declarante do SPED). Por padrao
  a tabela e lida de:
      C:\\Users\\jose.gleuton\\@ClaudePasta\\@Parametros\\PostosR7.csv
  Se nao existir nesse caminho, o programa pede o arquivo numa janela.

A descricao do produto vem do cadastro de itens (registro 0200, por COD_ITEM);
se nao houver, usa a descricao complementar do proprio C170.

Saidas (na pasta escolhida):
  - notas_entrada.csv       -> uma linha por nota selecionada (deduplicada)
  - itens_entrada.csv       -> uma linha por item (produto) das notas selecionadas
  - entrada_fornecedores.xlsx -> guia "Notas_Entrada" + guia "Itens_Entrada"
  - log_processamento.csv   -> uma linha por arquivo processado (auditoria)

Conferencia (a soma dos itens tem que bater com o total da nota):
  Para cada nota soma-se o VL_ITEM dos C170 e compara-se com o VL_DOC do C100.
    Diferenca = Valor Total da Nota - Soma dos Itens
    Confere   = OK / DIVERGE / SEM ITENS

Regras:
  - Considera apenas IND_OPER=0 (entrada) e COD_MOD=55 (NF-e)
  - Ignora notas canceladas/denegadas/inutilizadas (COD_SIT 02, 03, 04, 05)
  - Casa o fornecedor pela RAIZ do CNPJ (8 digitos) -> pega todas as filiais
  - Deduplica por (CNPJ declarante + chave NF-e), mantendo o arquivo mais
    recente; os itens acompanham a versao escolhida da nota
  - Nunca para por causa de um arquivo com problema: registra no log e continua

Uso:
    python sped_para_csv_forn_Item.py
        -> janelas para escolher SPEDs, tabela de fornecedores e pasta de saida
    python sped_para_csv_forn_Item.py C:/speds C:/tabela/Fornecedores.csv
    python sped_para_csv_forn_Item.py C:/speds C:/tabela/Fornecedores.csv C:/saida
    python sped_para_csv_forn_Item.py C:/speds C:/Forn.csv C:/saida C:/PostosR7.csv

Para gerar o XLSX e necessario o pacote openpyxl:
    pip install openpyxl
"""

__version__ = "2.1.0"

import csv
import io
import os
import re
import sys
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

CODSIT_EXCLUIR = {"02", "03", "04", "05"}  # canceladas, denegadas, inutilizadas

# Caminho padrao da tabela de postos (cruza pelo CNPJ da Sociedade)
POSTOS_PADRAO = r"C:\Users\jose.gleuton\@ClaudePasta\@Parametros\PostosR7.csv"


def so_digitos(s):
    return re.sub(r"\D", "", s or "")


def fmt_cnpj(c):
    c = so_digitos(c)
    if len(c) == 14:
        return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"
    if len(c) == 11:
        return f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}"
    return c


def fmt_data(d):
    d = (d or "").strip()
    return f"{d[0:2]}/{d[2:4]}/{d[4:8]}" if len(d) == 8 else d


def parse_num(v):
    """'19040,8' / '53855' / '1.234,56' -> float. Vazio/invalido -> 0.0."""
    v = (v or "").strip()
    if not v:
        return 0.0
    try:
        return float(v.replace(".", "").replace(",", ".")) if "," in v else float(v)
    except ValueError:
        return 0.0


def fmt_valor(f):
    """float -> '19040,80' (decimal com virgula, 2 casas)."""
    return f"{f:.2f}".replace(".", ",")


def fmt_qtd(f):
    """float -> quantidade com virgula, sem zeros a direita desnecessarios."""
    s = f"{f:.4f}".rstrip("0").rstrip(".")
    return s.replace(".", ",") if s else "0"


def _ler_csv(caminho):
    """Le um CSV (; ou ,) detectando codificacao e separador. -> (cabecalho, linhas)."""
    p = Path(caminho)
    raw = p.read_bytes()
    try:
        txt = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        txt = raw.decode("latin-1")
    delim = ";" if txt.count(";") >= txt.count(",") else ","
    linhas = list(csv.reader(io.StringIO(txt), delimiter=delim))
    if not linhas:
        return [], []
    return linhas[0], linhas[1:]


def carregar_fornecedores(caminho):
    """Le a tabela de fornecedores e devolve {raiz_cnpj(8 dig): apelido}."""
    raizes = {}
    cab, linhas = _ler_csv(caminho)
    if not cab:
        return raizes
    cab = [c.strip().lower() for c in cab]

    def idx(*nomes):
        for n in nomes:
            for i, c in enumerate(cab):
                if n in c:
                    return i
        return -1

    i_base = idx("cnpjbase", "cnpj base", "raiz")
    i_cnpj = idx("cnpj")
    i_ape = idx("apelido", "razao", "razão", "nome")

    for ln in linhas:
        if not ln:
            continue
        raiz = ""
        if 0 <= i_base < len(ln):
            raiz = so_digitos(ln[i_base])[:8]
        if len(raiz) < 8 and 0 <= i_cnpj < len(ln):
            raiz = so_digitos(ln[i_cnpj])[:8]
        if len(raiz) == 8:
            ape = ln[i_ape].strip() if 0 <= i_ape < len(ln) else ""
            raizes[raiz] = ape
    return raizes


def carregar_postos(caminho):
    """Le a tabela de postos e devolve {cnpj(14 dig): (cod, posto)}."""
    postos = {}
    if not caminho or not Path(caminho).exists():
        return postos
    cab, linhas = _ler_csv(caminho)
    if not cab:
        return postos
    cab = [c.strip().lower() for c in cab]

    def idx(*nomes):
        for n in nomes:
            for i, c in enumerate(cab):
                if n in c:
                    return i
        return -1

    i_cod = idx("cod", "codigo", "código")
    i_posto = idx("posto", "nome", "apelido")
    i_cnpj = idx("cnpj")
    if i_cnpj < 0:
        return postos

    for ln in linhas:
        if not ln or i_cnpj >= len(ln):
            continue
        cnpj = so_digitos(ln[i_cnpj])
        if len(cnpj) != 14:
            continue
        cod = ln[i_cod].strip() if 0 <= i_cod < len(ln) else ""
        posto = ln[i_posto].strip() if 0 <= i_posto < len(ln) else ""
        postos[cnpj] = (cod, posto)
    return postos


def processar_arquivo(args):
    """Processa um arquivo SPED. Roda em processo separado (worker).

    'args' = (caminho_str, raizes_selecionadas)
    Retorna apenas as notas dos fornecedores selecionados, ja com os itens.
    """
    caminho_str, raizes = args
    p = Path(caminho_str)
    res = {
        "arquivo": str(p),
        "mtime": 0.0,
        "empresa": "",
        "cnpj": "",
        "periodo": "",
        "notas": [],
        "status": "",
    }
    try:
        res["mtime"] = p.stat().st_mtime
        raw = p.read_bytes()
        try:
            texto = raw.decode("utf-8")
        except UnicodeDecodeError:
            texto = raw.decode("latin-1")

        if not texto.startswith("|0000|"):
            res["status"] = "ignorado: nao e arquivo SPED"
            return res

        parts = {}      # COD_PART -> (nome, doc) | registro 0150
        produtos = {}   # COD_ITEM -> descricao   | registro 0200
        nota_atual = None  # nota selecionada cujos C170 estamos coletando

        for ln in texto.splitlines():
            if not ln.startswith("|"):
                continue
            f = ln.split("|")
            reg = f[1]

            if reg == "0000":
                # |0000|COD_VER|COD_FIN|DT_INI|DT_FIN|NOME|CNPJ|...
                res["empresa"] = f[6].strip()
                res["cnpj"] = f[7].strip()
                res["periodo"] = f"{fmt_data(f[4])} a {fmt_data(f[5])}"

            elif reg == "0150":
                # |0150|COD_PART|NOME|COD_PAIS|CNPJ|CPF|IE|...
                if len(f) > 6:
                    parts[f[2].strip()] = (
                        f[3].strip(),
                        so_digitos(f[5]) or so_digitos(f[6]),
                    )

            elif reg == "0200":
                # |0200|COD_ITEM|DESCR_ITEM|...
                if len(f) > 3:
                    produtos[f[2].strip()] = f[3].strip()

            elif reg == "C100":
                # |C100|IND_OPER|IND_EMIT|COD_PART|COD_MOD|COD_SIT|SER|NUM|CHV|DT_DOC|DT_E_S|VL_DOC|
                nota_atual = None  # encerra a nota anterior
                if len(f) < 13:
                    continue
                if f[2].strip() != "0" or f[5].strip() != "55":
                    continue
                if f[6].strip() in CODSIT_EXCLUIR:
                    continue
                chave = f[9].strip()
                if len(chave) != 44:
                    continue
                cod_part = f[4].strip()
                nome_forn, doc_forn = parts.get(cod_part, ("", ""))
                raiz = doc_forn[:8] if len(doc_forn) == 14 else ""
                if raiz not in raizes:
                    continue  # fornecedor nao selecionado
                nota_atual = {
                    "empresa": res["empresa"],
                    "cnpj_decl": res["cnpj"],
                    "cnpj_forn": doc_forn,
                    "nome_forn": nome_forn,
                    "dt_doc": f[10].strip(),
                    "chave": chave,
                    "vl_doc": parse_num(f[12]),
                    "itens": [],
                }
                res["notas"].append(nota_atual)

            elif reg == "C170" and nota_atual is not None:
                # |C170|NUM_ITEM|COD_ITEM|DESCR_COMPL|QTD|UNID|VL_ITEM|VL_DESC|...
                if len(f) < 8:
                    continue
                cod = f[3].strip()
                descr = produtos.get(cod, "") or f[4].strip()
                nota_atual["itens"].append(
                    {
                        "num": f[2].strip(),
                        "cod": cod,
                        "descr": descr,
                        "qtd": parse_num(f[5]),
                        "unid": f[6].strip(),
                        "vl": parse_num(f[7]),
                    }
                )

        res["status"] = "ok"
    except Exception as e:
        res["status"] = f"ERRO: {e}"
    return res


def escolher_entradas():
    """Janelas para escolher a pasta dos SPEDs, a tabela de fornecedores,
    a pasta de saida e (se necessario) a tabela de postos."""
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        ent = input("Pasta-raiz com os arquivos SPED (.txt): ").strip(' "')
        tab = input("Arquivo da tabela de fornecedores (.csv): ").strip(' "')
        sai = input("Pasta para salvar as saidas (Enter = mesma dos SPEDs): ").strip(' "')
        pos = POSTOS_PADRAO if Path(POSTOS_PADRAO).exists() else \
            input("Arquivo PostosR7.csv (.csv): ").strip(' "')
        if not ent or not tab:
            print("Pasta/tabela nao informada. Operacao cancelada.")
            sys.exit(1)
        return Path(ent), Path(tab), Path(sai) if sai else Path(ent), pos

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    ent = filedialog.askdirectory(
        parent=root, title="Selecione a pasta-raiz com os arquivos SPED (.txt)")
    if not ent:
        print("Nenhuma pasta de SPED selecionada. Operacao cancelada.")
        root.destroy()
        sys.exit(1)

    tab = filedialog.askopenfilename(
        parent=root,
        title="Selecione a TABELA de fornecedores selecionados (.csv)",
        initialdir=ent,
        filetypes=[("CSV", "*.csv"), ("Todos os arquivos", "*.*")],
    )
    if not tab:
        print("Nenhuma tabela de fornecedores selecionada. Operacao cancelada.")
        root.destroy()
        sys.exit(1)

    sai = filedialog.askdirectory(
        parent=root, title="Selecione a pasta onde salvar as saidas", initialdir=ent)
    if not sai:
        print("Nenhuma pasta de saida selecionada. Operacao cancelada.")
        root.destroy()
        sys.exit(1)

    pos = POSTOS_PADRAO
    if not Path(pos).exists():
        pos = filedialog.askopenfilename(
            parent=root,
            title="Selecione a tabela PostosR7.csv",
            filetypes=[("CSV", "*.csv"), ("Todos os arquivos", "*.*")],
        )

    root.destroy()
    return Path(ent), Path(tab), Path(sai), pos


def conferencia(n):
    soma = sum(it["vl"] for it in n["itens"])
    dif = n["vl_doc"] - soma
    if not n["itens"]:
        conf = "SEM ITENS"
    elif abs(dif) < 0.01:
        conf = "OK"
    else:
        conf = "DIVERGE"
    return soma, dif, conf


CAB_NOTAS = [
    "Cod", "Posto",
    "Razao Social (SPED)", "CNPJ da Sociedade (SPED)", "CNPJ do Fornecedor",
    "Razao Social do Fornecedor", "Data de Emissao", "Chave XML",
    "Valor Total da Nota", "Soma dos Itens", "Diferenca", "Confere",
]
CAB_ITENS = [
    "Cod", "Posto",
    "Razao Social (SPED)", "CNPJ da Sociedade (SPED)", "CNPJ do Fornecedor",
    "Razao Social do Fornecedor", "Data de Emissao", "Chave XML",
    "Num Item", "Codigo do Item", "Descricao do Produto", "Quantidade",
    "Unidade", "Valor Total do Item", "Valor Total da Nota", "Confere Nota",
]


def gerar_xlsx(caminho_xlsx, notas, postos):
    """Gera o XLSX com as guias Notas_Entrada e Itens_Entrada.
    Retorna True se gerou; False se o openpyxl nao estiver instalado."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    wb = Workbook()
    ws_n = wb.active
    ws_n.title = "Notas_Entrada"
    ws_i = wb.create_sheet("Itens_Entrada")

    FMT = "#,##0.00"
    neg = Font(bold=True)

    def escreve_cabecalho(ws, cab):
        ws.append(cab)
        for c in ws[1]:
            c.font = neg
        ws.freeze_panes = "A2"

    escreve_cabecalho(ws_n, CAB_NOTAS)
    escreve_cabecalho(ws_i, CAB_ITENS)

    for n in notas:
        cod, posto = postos.get(so_digitos(n["cnpj_decl"]), ("", ""))
        soma, dif, conf = conferencia(n)
        ws_n.append([
            cod, posto, n["empresa"], fmt_cnpj(n["cnpj_decl"]),
            fmt_cnpj(n["cnpj_forn"]), n["nome_forn"], fmt_data(n["dt_doc"]),
            n["chave"], round(n["vl_doc"], 2), round(soma, 2), round(dif, 2), conf,
        ])
        for it in n["itens"]:
            ws_i.append([
                cod, posto, n["empresa"], fmt_cnpj(n["cnpj_decl"]),
                fmt_cnpj(n["cnpj_forn"]), n["nome_forn"], fmt_data(n["dt_doc"]),
                n["chave"], it["num"], it["cod"], it["descr"],
                round(it["qtd"], 4), it["unid"], round(it["vl"], 2),
                round(n["vl_doc"], 2), conf,
            ])

    # chave como texto e numeros formatados
    for col in ("H",):  # Chave XML (mesma coluna nas duas guias)
        for ws in (ws_n, ws_i):
            for cell in ws[col][1:]:
                cell.number_format = "@"
    for cell in ws_n["I"][1:] + ws_n["J"][1:] + ws_n["K"][1:]:
        cell.number_format = FMT
    for cell in ws_i["N"][1:] + ws_i["O"][1:]:
        cell.number_format = FMT

    # largura aproximada das colunas
    larguras_n = [6, 22, 34, 22, 22, 34, 13, 46, 16, 16, 14, 12]
    larguras_i = [6, 22, 34, 22, 22, 34, 13, 46, 9, 14, 40, 12, 8, 16, 16, 12]
    for i, w in enumerate(larguras_n, 1):
        ws_n.column_dimensions[get_column_letter(i)].width = w
    for i, w in enumerate(larguras_i, 1):
        ws_i.column_dimensions[get_column_letter(i)].width = w

    wb.save(caminho_xlsx)
    return True


def main():
    pos_path = POSTOS_PADRAO
    if len(sys.argv) > 4:
        pasta_ent, tabela, pasta_sai, pos_path = (
            Path(sys.argv[1]), Path(sys.argv[2]), Path(sys.argv[3]), sys.argv[4])
    elif len(sys.argv) > 3:
        pasta_ent, tabela, pasta_sai = Path(sys.argv[1]), Path(sys.argv[2]), Path(sys.argv[3])
    elif len(sys.argv) > 2:
        pasta_ent, tabela = Path(sys.argv[1]), Path(sys.argv[2])
        pasta_sai = pasta_ent
    else:
        pasta_ent, tabela, pasta_sai, pos_path = escolher_entradas()

    if not pasta_ent.exists():
        print(f"Pasta de SPED nao encontrada: {pasta_ent}")
        sys.exit(1)
    if not tabela.exists():
        print(f"Tabela de fornecedores nao encontrada: {tabela}")
        sys.exit(1)
    pasta_sai.mkdir(parents=True, exist_ok=True)

    raizes = carregar_fornecedores(tabela)
    if not raizes:
        print("A tabela de fornecedores nao tem nenhuma raiz de CNPJ valida.")
        print("Verifique se ha uma coluna 'CNPJBase' ou 'CNPJ'. Operacao cancelada.")
        sys.exit(1)

    postos = carregar_postos(pos_path)

    print("=" * 70)
    print("  Extrator de Notas e Itens de Entrada — SPED Fiscal -> CSV/XLSX")
    print("=" * 70)
    print(f"  SPEDs   : {pasta_ent}  (inclui subpastas)")
    print(f"  Tabela  : {tabela}")
    print(f"  Postos  : {pos_path}")
    print(f"  Saida   : {pasta_sai}")
    print(f"  Fornecedores selecionados: {len(raizes)} (por raiz de CNPJ)")
    print(f"  Postos carregados: {len(postos)}")
    if not postos:
        print("  [AVISO] Nenhum posto carregado — colunas Cod/Posto sairao vazias.")

    arquivos = sorted(str(p) for p in pasta_ent.rglob("*.txt"))
    print(f"  Arquivos .txt encontrados: {len(arquivos)}")
    print("=" * 70 + "\n")
    if not arquivos:
        print("Nenhum arquivo .txt encontrado.")
        sys.exit(0)

    t0 = time.time()
    workers = max(1, (os.cpu_count() or 2) - 1)
    print(f"Processando com {workers} processos em paralelo...\n")

    melhores = {}  # (cnpj declarante, chave) -> (mtime, nota_com_itens)
    log = []
    feitos = 0
    total = len(arquivos)
    larg = len(str(total))

    with ProcessPoolExecutor(max_workers=workers) as pool:
        futuros = {pool.submit(processar_arquivo, (a, raizes)): a for a in arquivos}
        for fut in as_completed(futuros):
            r = fut.result()
            n_itens = sum(len(n["itens"]) for n in r["notas"])
            log.append(
                [
                    Path(r["arquivo"]).name,
                    r["empresa"],
                    fmt_cnpj(r["cnpj"]),
                    r["periodo"],
                    len(r["notas"]),
                    n_itens,
                    r["status"],
                ]
            )
            for nota in r["notas"]:
                ch = (r["cnpj"], nota["chave"])
                atual = melhores.get(ch)
                if atual is None or r["mtime"] > atual[0]:
                    melhores[ch] = (r["mtime"], nota)
            feitos += 1

            # --- log de processamento ao vivo: uma linha por arquivo ---
            status = r["status"]
            if status == "ok":
                marca = "OK   "
            elif status.startswith("ignorado"):
                marca = "PULA "
            else:  # ERRO
                marca = "ERRO "
            pct = 100 * feitos / total
            empresa = (r["empresa"] or "-")[:35]
            nome_arq = Path(r["arquivo"]).name
            linha_log = (
                f"  [{feitos:>{larg}}/{total}] {pct:5.1f}%  {marca} "
                f"{empresa:<35}  notas: {len(r['notas']):>4}  itens: {n_itens:>5}  {nome_arq}"
            )
            if not status.startswith(("ok", "ignorado")):
                linha_log += f"  -> {status}"
            print(linha_log, flush=True)

    notas = [v[1] for v in melhores.values()]
    # ordena por posto (cod), empresa, data (AAAA-MM-DD), fornecedor
    def chave_ordem(n):
        cod, _ = postos.get(so_digitos(n["cnpj_decl"]), ("", ""))
        return (cod, n["empresa"],
                n["dt_doc"][4:8] + n["dt_doc"][2:4] + n["dt_doc"][0:2],
                n["nome_forn"])
    notas.sort(key=chave_ordem)

    # ---- notas_entrada.csv ----
    csv_notas = pasta_sai / "notas_entrada.csv"
    with open(csv_notas, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(CAB_NOTAS)
        for n in notas:
            cod, posto = postos.get(so_digitos(n["cnpj_decl"]), ("", ""))
            soma, dif, conf = conferencia(n)
            w.writerow(
                [
                    cod, posto,
                    n["empresa"],
                    fmt_cnpj(n["cnpj_decl"]),
                    fmt_cnpj(n["cnpj_forn"]),
                    n["nome_forn"],
                    fmt_data(n["dt_doc"]),
                    f'="{n["chave"]}"',  # protege a chave no Excel (CSV)
                    fmt_valor(n["vl_doc"]),
                    fmt_valor(soma),
                    fmt_valor(dif),
                    conf,
                ]
            )

    # ---- itens_entrada.csv ----
    csv_itens = pasta_sai / "itens_entrada.csv"
    with open(csv_itens, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(CAB_ITENS)
        for n in notas:
            cod, posto = postos.get(so_digitos(n["cnpj_decl"]), ("", ""))
            soma, dif, conf = conferencia(n)
            chave_xls = f'="{n["chave"]}"'
            for it in n["itens"]:
                w.writerow(
                    [
                        cod, posto,
                        n["empresa"],
                        fmt_cnpj(n["cnpj_decl"]),
                        fmt_cnpj(n["cnpj_forn"]),
                        n["nome_forn"],
                        fmt_data(n["dt_doc"]),
                        chave_xls,
                        it["num"],
                        it["cod"],
                        it["descr"],
                        fmt_qtd(it["qtd"]),
                        it["unid"],
                        fmt_valor(it["vl"]),
                        fmt_valor(n["vl_doc"]),
                        conf,
                    ]
                )

    # ---- log_processamento.csv ----
    csv_log = pasta_sai / "log_processamento.csv"
    log.sort(key=lambda x: x[0])
    with open(csv_log, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Arquivo", "Razao Social", "CNPJ", "Periodo",
                    "Notas Selecionadas", "Itens", "Status"])
        w.writerows(log)

    # ---- XLSX (2 guias) ----
    xlsx_path = pasta_sai / "entrada_fornecedores.xlsx"
    xlsx_ok = gerar_xlsx(xlsx_path, notas, postos)

    erros = sum(1 for l in log if str(l[6]).startswith("ERRO"))
    ignorados = sum(1 for l in log if str(l[6]).startswith("ignorado"))
    total_itens = sum(len(n["itens"]) for n in notas)
    n_ok = n_div = n_sem = 0
    for n in notas:
        _, _, conf = conferencia(n)
        if conf == "OK":
            n_ok += 1
        elif conf == "DIVERGE":
            n_div += 1
        else:
            n_sem += 1

    print("\n" + "=" * 70)
    print(f"  Concluido em {time.time() - t0:.1f}s")
    print(f"  Arquivos OK: {len(log) - erros - ignorados}   "
          f"Ignorados: {ignorados}   Erros: {erros}")
    print(f"  Notas selecionadas (apos dedup): {len(notas)}   Itens: {total_itens}")
    print(f"  Conferencia -> OK: {n_ok}   DIVERGE: {n_div}   SEM ITENS: {n_sem}")
    print(f"  -> {csv_notas}")
    print(f"  -> {csv_itens}")
    print(f"  -> {csv_log}")
    if xlsx_ok:
        print(f"  -> {xlsx_path}  (guias Notas_Entrada e Itens_Entrada)")
    else:
        print("  [AVISO] XLSX nao gerado: instale o openpyxl com  pip install openpyxl")
    if n_div:
        print(f"  [ATENCAO] {n_div} nota(s) com soma de itens diferente do total "
              f"(ver coluna 'Confere' = DIVERGE).")
    print("=" * 70)


if __name__ == "__main__":
    import multiprocessing

    multiprocessing.freeze_support()
    main()
